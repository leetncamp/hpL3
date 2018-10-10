#!/usr/bin/env python
from openpyxl import load_workbook
from argparse import ArgumentParser
import glob
import os, sys
from pdb import set_trace as debug
import ipaddress



parser = ArgumentParser(description="Add L3 to 9.32 Report")
parser.add_argument("files", nargs="*", help="Supply both L3 file and 9.32 file as Excel document on the command line")

subnet_header = "Cidr" #The networks
L3Override_header = "L3Override" #The person responsible.

ipaddress_header = "IP Addresses" #The column in 9.32 that contains the ipaddresses
LSM_header = "LSM"  #The column in the 9.32 spreadsheet after which I insert the results

ns = parser.parse_args()
print("Loading files...")
if ns.files:
    if "l3" in ns.files[0].lower() and "9.32" in ns.files[1].lower():
        ls_file = ns.files[0]
        n_file = ns.files[1]
    elif "l3" in ns.files[1] and "9.32" in ns.files[0].lower():
        ls_file = ns.files[1]
        n_file = ns.files[0]
        
else:
    #We didn't pass any files on the command line. Figure it out. 
    print(os.path.dirname(__file__))
    #os.chdir(os.path.dirname(__file__))
    files = [fn for fn in glob.glob("*.xlsx") if not fn.startswith("~")]
    l3files = [fn for fn in files if "l3" in fn.lower()]
    nfiles = [fn for fn in files if "9.32" in fn.lower()]
    exit = False
    if len(l3files) != 1:
        print("Too many or too few l3 files")
        exit = True
    if len(nfiles) != 1:
        print("Too many or too few 9.32 files")
        exit = True
    if exit:
        sys.exit(1)

    ls_file = l3files[0]
    n_file = nfiles[0]
l3wb = load_workbook(ls_file)
nwb = load_workbook(n_file)

l3ws = l3wb.active
nws = nwb.active


l3_headers = [i.value for i in l3ws[1]]
n_headers = [i.value for i in nws[1]]

subnet_col = l3_headers.index(subnet_header)
override_col = l3_headers.index(L3Override_header)
ip_col = n_headers.index(ipaddress_header)
lsm_col = n_headers.index(LSM_header) + 2


if "L3" not in n_headers:
    print("Inserting column...")
    nws.insert_cols(lsm_col)
    nws.cell(row=1, column=lsm_col).value = "L3"

l3 = {}

def calculate(iplookup):
    #Get all networks that match on the first 2 triplets. 

    first_two = ".".join(iplookup.split(".")[:2])
    networks = sorted([i for i in l3.keys() if i and i.startswith(first_two)])
    if not networks:
        print("No network found in L3 that matches the first 2 numbers. This isn't supposed to happen. Exiting!!!")
    result = "No L3 Found"
    for network in networks:
        try:
            if ipaddress.ip_address(iplookup) in ipaddress.ip_network(network, strict=False):
                result = l3.get(network)
                return(result)
        except Exception as e:
            result = "Error"
            return(e.message)

    return(result)



results = []
print("Calculating networks...")
for row in l3ws.iter_rows(min_row=2):
    l3[row[subnet_col].value] = row[override_col].value


for row in nws.iter_rows(min_row=2):
    iplookup = row[ip_col].value.split(",")[0]
    result = calculate(iplookup)
    results.append(result)
    row[lsm_col-1].value = result

newfile = os.path.join(os.getcwd(), "results", "{0}_L3.xlsx".format(os.path.splitext(n_file)[0]))
print("Saving...{0}".format(newfile))

nwb.save(newfile)
if os.uname[0] ==  "Darwin":
    os.system("open {0}".format(newfile))




