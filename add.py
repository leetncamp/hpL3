#!/usr/bin/env python
from openpyxl import load_workbook
from argparse import ArgumentParser
import glob
import os
from pdb import set_trace as debug
import ipaddress


parser = ArgumentParser(description="Add L3 to 9.32 Report")
parser.add_argument("files", nargs="*", help="Supply both L3 file and 9.32 file as Excel document on the command line")

subnet_header = "Cidr" #The networks
L3Override_header = "L3Override" #The person responsible.

ipaddress_header = "IP Addresses" #The column in 9.32 that contains the ipaddresses
LSM_header = "LSM"  #The column in the 9.32 spreadsheet after which I insert the results

ns = parser.parse_args()

if ns.files:
    if "l3" in ns.files[0].lower() and "9.32" in ns.files[1].lower():
        lswb = load_workbook(ns.files[0])
        nwb = load_workbook(ns.files[1])
    elif "l3" in ns.files[1] and "9.32" in ns.files[0].lower():
        l3wb = load_workbook(ns.files[1])
        nwb = load_workbook(ns.files[0])
else:
    #We didn't pass any files on the command line. Figure it out. 
    os.chdir(os.path.dirname(__file__))
    files = [fn for fn in glob.glob("*.xlsx") if not fn.startswith("~")]
    l3files = [fn for fn in files if "l3" in fn.lower()]
    nfiles = [fn for fn in files if "9.32" in fn.lower()]
    exit = False
    if len(l3files) != 1:
        print("Too many or too few l3 files")
        exit = True
    if len(nfiles) != 1:
        print("Too many or too few 9.32 files")
        exit = True
    if exit:
        sys.exit(1)
    l3wb = load_workbook(l3files[0]).active
    nwb = load_workbook(nfiles[0]).active


l3_headers = [i.value for i in l3wb[1]]
n_headers = [i.value for i in nwb[1]]

subnet_col = l3_headers.index(subnet_header)
override_col = l3_headers.index(L3Override_header)
ip_col = n_headers.index(ipaddress_header)
lsm_col = n_headers.index(LSM_header)

l3 = {}

def calculate(iplookup):
    #Get all networks that match on the first 2 triplets. 
    first_two = ".".join(iplookup.split(".")[:2])
    networks = sorted([i for i in l3.keys() if i and i.startswith(first_two)])
    result = None
    for network in networks:
        try:
            if ipaddress.ip_address(iplookup) in ipaddress.ip_network(network):
                result = l3.get(network)
                return(result)
        except:
            result = "Error"
            return(result)

    return(result)



results = []
print("Working...")
for row in l3wb.iter_rows(min_row=2):
    l3[row[subnet_col].value] = row[override_col].value

iplookups = []
for row in nwb.iter_rows(min_row=2):
    iplookup = row[ip_col].value
    result = calculate(iplookup)
    iplookups.append(iplookup)
    results.append(result)

print results

print("multi")
p = multiprocessing.Pool()

results = p.map(calculate, iplookup)
print results



